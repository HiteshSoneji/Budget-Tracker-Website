from django.shortcuts import render, redirect
from django.contrib.auth.models import User, auth
from django.contrib import messages
from .models import Ind_User
from django.core.mail import EmailMessage
import smtplib
import xlsxwriter
from django.core.files.storage import FileSystemStorage, default_storage
import os
from django.conf import settings
from django.http import JsonResponse
from openpyxl import Workbook
from openpyxl import load_workbook
from plotly.offline import plot
from plotly.graph_objs import Scatter

def register_view(request):
    if request.method == 'POST':
        f_name = request.POST['f_name']
        l_name = request.POST['l_name']
        username = request.POST['username']
        email = request.POST['email']
        password1 = request.POST['password1']
        password2 = request.POST['password2']
        if password1 == password2:
            if User.objects.filter(username=username).exists():
                print("Username Exists")
            elif User.objects.filter(email=email).exists():
                print("email Taken")
            else:
                user = User.objects.create_user(username=username, password=password1, email=email, first_name=f_name, last_name = l_name)
                user.save();
                s = Ind_User(username=username, l_name=l_name, f_name=f_name, password=password1, email=email)
                s.save();
                headers = ['loans', 'utility bills', 'insurance', 'entertainment', 'groceries', 'transportation', 'retirement fund', 'emergency fund', 'childcare and school costs', 'clothing', 'maintainance']
                wb = Workbook()
                ws = wb.active
                ws.title='expenses'
                ws.append(headers)
                u = username
                u = str(u)
                wb.save(settings.MEDIA_ROOT+r'\uploads\Indiv_'+u+'_Data.xlsx')
                fn = (r'uploads\Indiv_'+u+'_Data.xlsx')
                s.e_file = fn
                s.save()
        else:
            print("Password does not match")
        return redirect('/')
    else:
        return render(request, 'accounts/register.html')

def login_view(request):
    if request.method=='POST':
        username = request.POST['username']
        password1 = request.POST['password1']

        user = auth.authenticate(username=username, password = password1)

        if user is not None:
            auth.login(request, user)
            return redirect('/')

        else:
            print("Error")
            return redirect('')
    return render(request, 'accounts/login.html')


def home(request):
    if request.user.is_authenticated:
        u = request.user.username
        u = str(u)
        fn = (settings.MEDIA_ROOT+r'\uploads\Indiv_'+u+'_Data.xlsx')
        fn = str(fn)
        print(fn)
        workbook = load_workbook(fn)
        # ws = workbook.get_sheet_by_name('expenses')
        ws = workbook.active
        a1 = ws["A1"]
        # print(a1.value)
        if request.method == 'POST':
            l = request.POST['myselection']
            l = str(l)
            l = l.lower()
            # print(l)
            bal = request.POST['exp']
            # print(bal)
            row = 1
            column = 1
            for i in range(1,12):
                ref = ws.cell(row = row, column = i)
                ref_value = ref.value
                ref_value = str(ref_value)
                ref_value = ref_value.lower()
                # print(ref_value)
                if ref_value == l:
                    ws.cell(row = row + 1, column = i, value = bal)
                    workbook.save(settings.MEDIA_ROOT+r'\uploads\Indiv_'+u+'_Data.xlsx')
                    break
                else:
                    continue

    return render(request, 'accounts/home.html')

def logout_view(request):
    auth.logout(request)
    return render(request, 'accounts/login.html')

def graph_view(request):
    x_data = [0,1,2,3]
    y_data = [x**2 for x in x_data]
    plot_div = plot([Scatter(x=x_data, y=y_data,
                             mode='lines', name='test',
                             opacity=0.8, marker_color='green')],
                    output_type='div')
    return render(request, "accounts/home_graph.html", context={'plot_div': plot_div})